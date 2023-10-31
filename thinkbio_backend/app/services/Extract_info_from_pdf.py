# service/extract_info_from_pdf
import PyPDF2
import re
from datetime import datetime
from tqdm import tqdm
import openpyxl


def extract_text_from_pdf(pdf_file_path):
    text = ""
    with open(pdf_file_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        num_pages = len(pdf_reader.pages)
        for page_num in tqdm(range(num_pages), desc="Extraction du texte"):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            page_text = remove_occurrences_from_text(page_text, "243807A")
            page_text = remove_Page(page_text)
            page_text = remove_custom_pattern(page_text, r'Edité.{13}')
            page_text = remove_custom_pattern(page_text, r'Tiers RelaDate RelatiType dLibelle Relation')
            text += page_text
    return text

def remove_occurrences_from_text(text, pattern):
    return text.replace(pattern, '')

def remove_Page(text):
    return re.sub(r'Page\d*', '', text)

def remove_custom_pattern(text, pattern):
    return re.sub(pattern, '', text)

def extract_values(extracted_text):
    # Définition des REGEX
    date_intervention_pattern = r'(\d{2}/\d{2}/\d{4})\s*?RDV'
    boues_pattern = r'Boues\s*:\s*(\d+)\s*cm'
    croutes_pattern = r'Croutes\s*:\s*(\d+)?\s*cm\n?'
    pressure_pattern = r'Pression\s*:\s*fermés\s*:\s*([\d.]+)?\s*Ouverts\s*:\s*([\d.]+)?\s*Maximale\s*:\s*([\d.]+)?\s*Après décolmatage\s*:?\s*?([\d.]+)?\s*?\n?'
    bullage_pattern = r'Bullage\s*:\s*((?:(?:Fin|Faible|Grossier|Inexistant)(?=\s*\[X\]))|(?:Fin|Faible|Grossier|Inexistant))'
    condenstation_pattern = r'Condensation\s*:\s*((?:(?:Non|Peu|Beaucoup|Persistant)(?=\s*\[X\]))|(?:Non|Peu|Beaucoup|Persistant))'
    typeIntervention_pattern = r'\n?RDV\n?(.*?)\nLieu d\'intervention :'
    Compresseur_pattern = r'Compresseur\s*:\s*([^\n:]+)?\s*Dates\s*:\s*([^:]+)?\s*Pompe'
    Pompe_pattern = r'Pompe\s*:\s*([^\n:]+)?\s*Dates\s*:\s*([^:]+)?\s*\n'
    Membrane_pattern = r'Membranes\s*:\s*([^\n:]+)?\s*Dates\s*:\s*([^:]+)?\s*Marteau'
    Marteau_pattern = r'Marteau\s*:\s*([^\n:]+)?\s*Dates\s*:\s*([^:]+)?\s*\n'
    Diffuseur_pattern = r'Diffuseur\s*:\s*([^\n:]+)?\s*Dates\s*:\s*([^:]+)?\s*Filtre'
    
    # Définition des textes de capture
    date_intervention_values = re.findall(date_intervention_pattern, extracted_text)
    boues_values = re.findall(boues_pattern, extracted_text)
    croutes_values = re.findall(croutes_pattern, extracted_text)
    pressure_matches = re.finditer(pressure_pattern, extracted_text)
    bullage_values = re.findall(bullage_pattern, extracted_text)
    condenstation_values = re.findall(condenstation_pattern, extracted_text)
    type_intervention = re.search(typeIntervention_pattern, extracted_text, re.DOTALL)
    Compresseur_information = re.search(Compresseur_pattern, extracted_text)
    Pompe_information = re.search(Pompe_pattern, extracted_text)
    Membrane_information = re.search(Membrane_pattern, extracted_text)
    Marteau_information = re.search(Marteau_pattern, extracted_text)
    Diffuseur_information = re.search(Diffuseur_pattern, extracted_text)

    pressure_values = []

    for match in pressure_matches:
        fermes_val, ouverts_val, maximale_val, apres_decolmatage_val = match.groups()
        pressure_values.append({
            "Fermé": fermes_val if fermes_val else None,
            "Ouvert": ouverts_val if ouverts_val else None,
            "Maximale": maximale_val if maximale_val else None,
            "Après décolmatage": apres_decolmatage_val if apres_decolmatage_val else None
        })

    boues_values = [int(value) if value else None for value in boues_values]
    croutes_values = [int(value) if value else None for value in croutes_values]

    date_values = [datetime.strptime(value, "%d/%m/%Y").strftime("%d/%m/%Y") for value in date_intervention_values]

    boues_list = boues_values  # Mettre les valeurs des boues dans une liste

    croute_list = croutes_values  # Mettre les valeurs des croutes dans une liste

    if type_intervention:
        captured_text = type_intervention.group(1)
    else:
        captured_text = ""  # Ou toute autre valeur par défaut

    # Extraire les informations du Compresseur
    Compresseur_information_dict = {}
    if Compresseur_information:
        compresseur_text, dates_text = Compresseur_information.groups()
        Compresseur_information_dict = {
            "Compresseur": compresseur_text.strip() if compresseur_text else "",
            "Dates": dates_text.strip() if dates_text else ""
        }
        
    # Extraire les informations de la pompe
    Pompe_information_dict = {}
    if Pompe_information:
        pompe_text, datespompe_text = Pompe_information.groups()
        Pompe_information_dict = {
            "Pompe": pompe_text.strip() if pompe_text else "",
            "Dates": datespompe_text.strip() if datespompe_text else ""
        }
    
    # Extraire les informations des membranes
    Membrane_information_dict = {}
    if Membrane_information:
        membrane_texte, datesmembrane_text = Membrane_information.groups()
        Membrane_information_dict = {
            "Membrane" : membrane_texte.strip() if membrane_texte else "",
            "Dates": datesmembrane_text.strip() if datesmembrane_text else ""
        }
        
    # Extraire les informations des marteaux
    Marteau_information_dict = {}
    if Marteau_information:
        Marteau_texte, datesMarteau_text = Marteau_information.groups()
        Marteau_information_dict = {
            "Marteau" : Marteau_texte.strip() if Marteau_texte else "",
            "Dates": datesMarteau_text.strip() if datesMarteau_text else ""
        }

    # Extraire les informations des diffuseurs 
    Diffuseur_information_dict = {}
    if Diffuseur_information:
        Diffuseur_texte, datesDiffuseur_text = Diffuseur_information.groups()
        Diffuseur_information_dict = {
            "Diffuseur" : Diffuseur_texte.strip() if Diffuseur_texte else "",
            "Dates": datesDiffuseur_text.strip() if datesDiffuseur_text else ""
        }
     
    return boues_list, croute_list, date_values, pressure_values, bullage_values, condenstation_values, captured_text, Compresseur_information_dict, Pompe_information_dict, Membrane_information_dict, Marteau_information_dict, Diffuseur_information_dict


def search_patterns(text):
    patterns = [
        r'\b\dA\d{5}',
        r'\b\d{2}A\d{4}'
    ]

    results = []

    all_matches = []

    for pattern in patterns:
        matches = re.finditer(pattern, text)
        matches = list(matches)
        all_matches.extend(matches)

    all_matches.sort(key=lambda match: match.start())

    for i in range(len(all_matches) - 1):
        start, end = all_matches[i].span()
        next_start, _ = all_matches[i + 1].span()
        extracted_text = text[start + 7:next_start]
        boues_list, croute_list, date_values, pressure_values, bullage_values, condenstation_values, type_intervention, Compresseur_information, Pompe_information_dict, Membrane_information_dict, Marteau_information_dict, Diffuseur_information_dict = extract_values(extracted_text)

        result_dict = {
            "station": text[start:start + 7],
            "type_intervention": type_intervention,
            "extracted_text": extracted_text,
            "Boues": boues_list,
            "Croutes": croute_list,
            "Dates": date_values,
            "Pression": pressure_values,
            "bullage_quality": bullage_values,
            "condenstation_quality": condenstation_values,
            "Compresseur_information": Compresseur_information,
            "Pompe_information": Pompe_information_dict,
            "Membrane_information": Membrane_information_dict,
            "Marteau_information": Marteau_information_dict,
            "Diffuseur_information": Diffuseur_information_dict,
        }

        results.append(result_dict)

    return results

def create_excel_file(data):
    # Création du nouveau classeur
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Écriture des en-têtes
    headers = [
        "Station",
        "Date",
        "Type d'intervention",
        "Boues année dernière (cm)",
        "Boues année en cours (cm)",
        "Croutes année dernière (cm)",
        "Croutes année en cours (cm)",
        "Pression fermé année dernière",
        "Pression fermé année en cours",
        "Pression ouvert année dernière",
        "Pression ouvert année en cours",
        "Pression maximale année dernière",
        "Pression maximale année en cours",
        "Pression après décolmatage année dernière",
        "Pression après décolmatage année en cours",
        "Qualité du bullage l'année dernière",
        "Qualité du bullage année en cours",
        "Qualification condenstation année dernière",
        "Qualification condenstation année en cours",
        "Compresseur MODELE",
        "Date Changement COMPRESSEUR",
        "Pompe MODELE",
        "Date Changement POMPE",
        "Membrane MODELE",
        "Date Changement Membrane",
        "Marteau MODELE",
        "Date Changement Marteau"
        "Marteau Diffuseur",
        "Diffuseur Modele",
        "Date Changement Diffuseur"
    ]

    for col, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col, value=header)

    # Écriture des données
    for row, record in enumerate(data, start=3):
        sheet.cell(row=row, column=1, value=record["station"])
        
        if record["Dates"]:
            sheet.cell(row=row, column=2, value=record["Dates"][0])
        else:
            sheet.cell(row=row, column=2, value="")  # Gérer le cas où la liste Dates est vide
        
        if record["type_intervention"]:
            sheet.cell(row=row, column=3, value=record["type_intervention"])
        else:
            sheet.cell(row=row, column=3, value="")  # Gérer le cas où la liste Dates est vide

        if "Boues" in record:
            boues_data = record["Boues"]
            if len(boues_data) >= 2:
                sheet.cell(row=row, column=4, value=boues_data[1])  # Boues année dernière
                sheet.cell(row=row, column=5, value=boues_data[0])  # Boues année en cours
            elif len(boues_data) == 1:
                sheet.cell(row=row, column=5, value=boues_data[0])  # Boues année en cours

        if "Croutes" in record:
            croutes_data = record["Croutes"]
            if len(croutes_data) >= 2:
                sheet.cell(row=row, column=6, value=croutes_data[1])  # Croutes année dernière
                sheet.cell(row=row, column=7, value=croutes_data[0])  # Croutes année en cours
            elif len(croutes_data) == 1:
                sheet.cell(row=row, column=7, value=croutes_data[0])  # Croutes année en cours
    
        if "Pression" in record:
            pression_data = record["Pression"]
            if pression_data:
                if len(pression_data) >= 2:
                    sheet.cell(row=row, column=8, value=pression_data[1]["Fermé"])  # Pression fermé année dernière
                    sheet.cell(row=row, column=9, value=pression_data[0]["Fermé"])  # Pression fermé année en cours

                    sheet.cell(row=row, column=10, value=pression_data[1]["Ouvert"])  # Pression ouvert année dernière
                    sheet.cell(row=row, column=11, value=pression_data[0]["Ouvert"])  # Pression ouvert année en cours

                    sheet.cell(row=row, column=12, value=pression_data[1]["Maximale"])  # Pression maximale année dernière
                    sheet.cell(row=row, column=13, value=pression_data[0]["Maximale"])  # Pression maximale année en cours

                    sheet.cell(row=row, column=14, value=pression_data[1]["Après décolmatage"])  # Pression après décolmatage année dernière
                    sheet.cell(row=row, column=15, value=pression_data[0]["Après décolmatage"])  # Pression après décolmatage année en cours
                elif len(pression_data) == 1:
                    sheet.cell(row=row, column=8, value=pression_data[0]["Fermé"])  # Pression fermé année en cours
                    sheet.cell(row=row, column=9, value="")  # Pas de pression fermé année dernière

                    sheet.cell(row=row, column=10, value=pression_data[0]["Ouvert"])  # Pression ouvert année en cours
                    sheet.cell(row=row, column=11, value="")  # Pas de pression ouvert année dernière

                    sheet.cell(row=row, column=12, value=pression_data[0]["Maximale"])  # Pression maximale année en cours
                    sheet.cell(row=row, column=13, value="")  # Pas de pression maximale année dernière

                    sheet.cell(row=row, column=14, value=pression_data[0]["Après décolmatage"])  # Pression après décolmatage année en cours
                    sheet.cell(row=row, column=15, value="")  # Pas de pression après décolmatage année dernière
        
        if "bullage_quality" in record:
            quality_values = record["bullage_quality"]
            if len(quality_values) >=2:
                sheet.cell(row=row, column=16, value=quality_values[1])
                sheet.cell(row=row, column=17, value=quality_values[0])
            elif len(quality_values) == 1:
                sheet.cell(row=row, column=16, value=quality_values[0])
        
        if "condenstation_quality" in record:
            condenstation_values = record["condenstation_quality"]
            if len(condenstation_values) >=2:
                sheet.cell(row=row, column=18, value=condenstation_values[1])
                sheet.cell(row=row, column=19, value=condenstation_values[0])
            elif len(condenstation_values) == 1:
                sheet.cell(row=row, column=18, value=condenstation_values[0])
                
        if "Compresseur_information" in record:
            compressor_info = record["Compresseur_information"]
            compresseur = compressor_info.get("Compresseur", "")
            dates = compressor_info.get("Dates", "")

            sheet.cell(row=row, column=20, value=compresseur)
            sheet.cell(row=row, column=21, value=dates)
        
        if "Pompe_information" in record:
            pompe_info = record["Pompe_information"]
            pompe = pompe_info.get("Pompe", "")
            datesPompe = pompe_info.get("Dates", "")
            
            sheet.cell(row=row, column=22, value=pompe)
            sheet.cell(row=row, column=23, value=datesPompe)
            
        if "Membrane_information" in record:
            menbrane_info = record["Membrane_information"]
            membrane = menbrane_info.get("Membrane", "")
            datesMembrane = menbrane_info.get("Dates", "")
            
            sheet.cell(row=row, column=24, value=membrane)
            sheet.cell(row=row, column=25, value=datesMembrane)
            
        if "Marteau_information" in record:
            marteau_info = record["Marteau_information"]
            Marteau = marteau_info.get("Marteau", "")
            datesMarteau = marteau_info.get("Dates", "")
            
            sheet.cell(row=row, column=26, value=Marteau)
            sheet.cell(row=row, column=27, value=datesMarteau)
            
        if "Diffuseur_information" in record:
            Diffuseur_info = record["Diffuseur_information"]
            Diffuseur = Diffuseur_info.get("Diffuseur", "")
            datesDiffuseur = Diffuseur_info.get("Dates", "")
            
            sheet.cell(row=row, column=28, value=Diffuseur)
            sheet.cell(row=row, column=29, value=datesDiffuseur)
        
        
    # Sauvegarde du fichier
    wb.save('./ressources/ouput/GRC/output.xlsx')

    return None